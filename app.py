import math
import os
from datetime import datetime
from functools import wraps
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from flask import (Flask, flash, redirect, render_template,
                   request, send_file, session, url_for)
from werkzeug.security import check_password_hash

from db import get_db

import json as _json

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'csupd-dev-secret-2024')
app.jinja_env.filters['from_json'] = _json.loads

# ── Константы ────────────────────────────────────────────────────────────────

ASSEMBLY_STAGES = [
    'Входной контроль', 'Разборка', 'Дефектация',
    'Сборка узла', 'Сборка двигателя',
    'Стендовые испытания', 'Выходной контроль',
]

ROLE_LEVEL = {'sborshik': 1, 'kladovshik': 2, 'master': 3, 'admin': 4}
ROLE_NAMES = {
    'sborshik':   'Сборщик',
    'kladovshik': 'Кладовщик',
    'master':     'Мастер',
    'admin':      'Администратор',
}

# ── Вспомогательные функции авторизации ──────────────────────────────────────

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.full_path))
        return f(*args, **kwargs)
    return wrapper


def role_required(min_role):
    """Пропускает только пользователей с уровнем роли >= min_role."""
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login', next=request.full_path))
            if ROLE_LEVEL.get(session.get('role'), 0) < ROLE_LEVEL.get(min_role, 99):
                flash('Недостаточно прав для этого действия.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return wrapper
    return decorator


@app.context_processor
def inject_user():
    """Вставляет current_user и ROLE_LEVEL во все шаблоны."""
    u = None
    if 'user_id' in session:
        u = {
            'id':        session['user_id'],
            'username':  session['username'],
            'role':      session['role'],
            'full_name': session['full_name'],
            'role_name': ROLE_NAMES.get(session['role'], session['role']),
            'level':     ROLE_LEVEL.get(session['role'], 0),
        }
    return {'current_user': u, 'ROLE_LEVEL': ROLE_LEVEL}


# ── Авторизация ───────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    next_url = request.args.get('next', '')

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        next_url = request.form.get('next', '')

        db = get_db()
        user = db.execute(
            'SELECT * FROM users WHERE username = ?', (username,)
        ).fetchone()
        db.close()

        if user and check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id']   = user['id']
            session['username']  = user['username']
            session['role']      = user['role']
            session['full_name'] = user['full_name'] or user['username']
            flash(f'Добро пожаловать, {session["full_name"]}!', 'success')
            return redirect(next_url or url_for('dashboard'))

        flash('Неверный логин или пароль.', 'danger')

    return render_template('login.html', next=next_url)


@app.route('/logout')
def logout():
    name = session.get('full_name', '')
    session.clear()
    flash(f'Вы вышли из системы{(", " + name) if name else ""}.', 'success')
    return redirect(url_for('login'))


# ── Дашборд ───────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    q = request.args.get('q', '').strip()
    db = get_db()

    sql = '''
        SELECT p.id, p.article, p.name, p.engine_type, p.unit,
               p.monthly_plan, p.red_threshold, p.yellow_threshold,
               s.quantity, sc.cell_code,
               CASE
                 WHEN s.quantity < p.red_threshold   THEN 'red'
                 WHEN s.quantity < p.yellow_threshold THEN 'yellow'
                 ELSE 'green'
               END AS status,
               CASE
                 WHEN s.quantity < p.red_threshold   THEN 1
                 WHEN s.quantity < p.yellow_threshold THEN 2
                 ELSE 3
               END AS sort_order
        FROM parts p
        JOIN stock s  ON s.part_id = p.id
        JOIN storage_cells sc ON sc.id = s.cell_id
    '''
    params = []
    if q:
        sql += ' WHERE (p.name LIKE ? OR p.article LIKE ?)'
        params = [f'%{q}%', f'%{q}%']
    sql += ' ORDER BY sort_order, p.name'

    parts = db.execute(sql, params).fetchall()

    red_count    = sum(1 for p in parts if p['status'] == 'red')
    yellow_count = sum(1 for p in parts if p['status'] == 'yellow')
    green_count  = sum(1 for p in parts if p['status'] == 'green')

    movements = db.execute('''
        SELECT m.operation_type, m.quantity, m.operator, m.created_at,
               p.name AS part_name, p.unit
        FROM movements m
        JOIN parts p ON m.part_id = p.id
        ORDER BY m.created_at DESC
        LIMIT 10
    ''').fetchall()

    db.close()
    return render_template('dashboard.html',
        parts=parts,
        red_count=red_count,
        yellow_count=yellow_count,
        green_count=green_count,
        movements=movements,
        q=q,
    )


# ── Расход ────────────────────────────────────────────────────────────────────

@app.route('/raskhod', methods=['GET', 'POST'])
@login_required
def raskhod():
    db = get_db()

    if request.method == 'POST':
        part_id        = request.form.get('part_id', '').strip()
        quantity_raw   = request.form.get('quantity', '').strip()
        engine_number  = request.form.get('engine_number', '').strip()
        assembly_stage = request.form.get('assembly_stage', '').strip()
        operator       = session.get('full_name') or session.get('username', '')

        errors = []
        if not part_id:       errors.append('Выберите деталь.')
        if not engine_number: errors.append('Укажите номер двигателя.')
        if not assembly_stage:errors.append('Выберите этап сборки.')
        if not quantity_raw or not quantity_raw.isdigit() or int(quantity_raw) < 1:
            errors.append('Количество должно быть целым числом больше нуля.')

        if errors:
            for e in errors:
                flash(e, 'danger')
        else:
            quantity = int(quantity_raw)
            stock = db.execute(
                'SELECT s.id, s.quantity, s.cell_id FROM stock s WHERE s.part_id = ?',
                (part_id,)
            ).fetchone()

            if not stock:
                flash('Деталь не найдена в остатках.', 'danger')
            elif quantity > stock['quantity']:
                flash(
                    f'Недостаточно на складе. '
                    f'Запрошено: {quantity}, в наличии: {stock["quantity"]} шт.',
                    'danger'
                )
            else:
                db.execute(
                    'UPDATE stock SET quantity = quantity - ?, updated_at = CURRENT_TIMESTAMP '
                    'WHERE id = ?',
                    (quantity, stock['id'])
                )
                db.execute(
                    'INSERT INTO movements '
                    '(part_id, cell_id, operation_type, quantity, '
                    'engine_number, assembly_stage, operator) '
                    'VALUES (?, ?, ?, ?, ?, ?, ?)',
                    (part_id, stock['cell_id'], 'расход',
                     quantity, engine_number, assembly_stage, operator)
                )
                db.commit()
                part = db.execute(
                    'SELECT name, article FROM parts WHERE id = ?', (part_id,)
                ).fetchone()
                flash(
                    f'Расход зафиксирован: {part["name"]} ({part["article"]}) — '
                    f'{quantity} шт. · двигатель {engine_number}',
                    'success'
                )
                db.close()
                return redirect(url_for('dashboard'))

    parts = db.execute('''
        SELECT p.id, p.article, p.name, p.unit, s.quantity
        FROM parts p
        JOIN stock s ON s.part_id = p.id
        ORDER BY p.name
    ''').fetchall()

    db.close()
    return render_template('raskhod.html',
        parts=parts,
        assembly_stages=ASSEMBLY_STAGES,
        form=request.form,
        current_operator=session.get('full_name') or session.get('username', ''),
    )


# ── Приход ────────────────────────────────────────────────────────────────────

@app.route('/prikhod', methods=['GET', 'POST'])
@role_required('kladovshik')
def prikhod():
    db = get_db()

    if request.method == 'POST':
        part_id      = request.form.get('part_id', '').strip()
        quantity_raw = request.form.get('quantity', '').strip()
        cell_id      = request.form.get('cell_id', '').strip()
        operator     = session.get('full_name') or session.get('username', '')
        notes        = request.form.get('notes', '').strip()

        errors = []
        if not part_id:  errors.append('Выберите деталь.')
        if not quantity_raw or not quantity_raw.isdigit() or int(quantity_raw) < 1:
            errors.append('Количество должно быть целым числом больше нуля.')

        if errors:
            for e in errors:
                flash(e, 'danger')
        else:
            quantity = int(quantity_raw)
            stock = db.execute(
                'SELECT s.id, s.quantity, s.cell_id FROM stock s WHERE s.part_id = ?',
                (part_id,)
            ).fetchone()

            target_cell_id = int(cell_id) if cell_id else (stock['cell_id'] if stock else None)

            if stock:
                db.execute(
                    'UPDATE stock SET quantity = quantity + ?, updated_at = CURRENT_TIMESTAMP '
                    'WHERE id = ?',
                    (quantity, stock['id'])
                )
            else:
                db.execute(
                    'INSERT INTO stock (part_id, cell_id, quantity) VALUES (?, ?, ?)',
                    (part_id, target_cell_id, quantity)
                )

            db.execute(
                'INSERT INTO movements '
                '(part_id, cell_id, operation_type, quantity, operator, notes) '
                'VALUES (?, ?, ?, ?, ?, ?)',
                (part_id, target_cell_id, 'приход', quantity, operator, notes or None)
            )
            db.commit()
            part = db.execute(
                'SELECT name, article FROM parts WHERE id = ?', (part_id,)
            ).fetchone()
            flash(
                f'Приход зафиксирован: {part["name"]} ({part["article"]}) — {quantity} шт.',
                'success'
            )
            db.close()
            return redirect(url_for('dashboard'))

    parts = db.execute('''
        SELECT p.id, p.article, p.name, p.unit,
               s.quantity, s.cell_id, sc.cell_code
        FROM parts p
        JOIN stock s  ON s.part_id = p.id
        JOIN storage_cells sc ON sc.id = s.cell_id
        ORDER BY p.name
    ''').fetchall()

    cells = db.execute(
        'SELECT id, cell_code FROM storage_cells ORDER BY cell_code'
    ).fetchall()

    selected_part_id = request.args.get('part_id', '')
    db.close()
    return render_template('prikhod.html',
        parts=parts,
        cells=cells,
        form=request.form,
        selected_part_id=selected_part_id,
        current_operator=session.get('full_name') or session.get('username', ''),
    )


# ── Заявки на пополнение ──────────────────────────────────────────────────────

@app.route('/zayavki')
@role_required('kladovshik')
def zayavki():
    status_filter = request.args.get('status', '')
    db = get_db()

    counts = {r['status']: r['cnt'] for r in db.execute(
        'SELECT status, COUNT(*) AS cnt FROM replenishment_requests GROUP BY status'
    ).fetchall()}

    sql = '''
        SELECT rr.id, rr.quantity, rr.status, rr.created_by,
               rr.created_at, rr.completed_at,
               p.name AS part_name, p.article, p.unit,
               p.monthly_plan, s.quantity AS stock_qty
        FROM replenishment_requests rr
        JOIN parts p ON rr.part_id = p.id
        JOIN stock s ON s.part_id = p.id
    '''
    params = []
    if status_filter:
        sql += ' WHERE rr.status = ?'
        params.append(status_filter)
    sql += ' ORDER BY rr.created_at DESC'

    requests_list = db.execute(sql, params).fetchall()
    db.close()

    return render_template('zayavki.html',
        requests_list=requests_list,
        counts=counts,
        status_filter=status_filter,
    )


@app.route('/zayavki/create', methods=['POST'])
@role_required('kladovshik')
def zayavki_create():
    part_id    = request.form.get('part_id', '').strip()
    quantity   = request.form.get('quantity', '').strip()
    created_by = request.form.get('created_by', '').strip()

    if not part_id or not quantity or not quantity.isdigit() or int(quantity) < 1:
        flash('Некорректные данные для создания заявки.', 'danger')
        return redirect(url_for('dashboard'))

    db = get_db()
    db.execute(
        'INSERT INTO replenishment_requests (part_id, quantity, status, created_by) '
        'VALUES (?, ?, ?, ?)',
        (part_id, int(quantity), 'Создана',
         created_by or session.get('full_name', 'Неизвестно'))
    )
    db.commit()
    part = db.execute('SELECT name, article FROM parts WHERE id = ?', (part_id,)).fetchone()
    db.close()

    flash(f'Заявка создана: {part["name"]} ({part["article"]}) — {quantity} шт.', 'success')
    return redirect(request.form.get('next', url_for('dashboard')))


@app.route('/zayavki/<int:req_id>/update', methods=['POST'])
@role_required('kladovshik')
def zayavki_update(req_id):
    new_status = request.form.get('status', '')
    if new_status not in ('В работе', 'Выполнена'):
        flash('Недопустимый статус.', 'danger')
        return redirect(url_for('zayavki'))

    db = get_db()
    if new_status == 'Выполнена':
        db.execute(
            'UPDATE replenishment_requests '
            'SET status = ?, completed_at = CURRENT_TIMESTAMP WHERE id = ?',
            (new_status, req_id)
        )
    else:
        db.execute(
            'UPDATE replenishment_requests SET status = ? WHERE id = ?',
            (new_status, req_id)
        )
    db.commit()
    db.close()

    flash(f'Статус заявки #{req_id} обновлён: {new_status}', 'success')
    return redirect(url_for('zayavki'))


# ── Журнал операций ───────────────────────────────────────────────────────────

@app.route('/journal')
@login_required
def journal():
    q         = request.args.get('q', '').strip()
    op_type   = request.args.get('op_type', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    page      = max(1, int(request.args.get('page', 1) or 1))
    per_page  = 50

    db = get_db()

    conditions = ['1=1']
    params     = []

    if q:
        conditions.append(
            '(p.name LIKE ? OR p.article LIKE ? OR m.operator LIKE ? OR m.engine_number LIKE ?)'
        )
        params += [f'%{q}%'] * 4
    if op_type in ('приход', 'расход'):
        conditions.append('m.operation_type = ?')
        params.append(op_type)
    if date_from:
        conditions.append('DATE(m.created_at) >= ?')
        params.append(date_from)
    if date_to:
        conditions.append('DATE(m.created_at) <= ?')
        params.append(date_to)

    where = ' AND '.join(conditions)

    total = db.execute(
        f'SELECT COUNT(*) FROM movements m JOIN parts p ON m.part_id = p.id WHERE {where}',
        params,
    ).fetchone()[0]

    movements = db.execute(f'''
        SELECT m.id, m.operation_type, m.quantity, m.engine_number,
               m.assembly_stage, m.operator, m.created_at, m.notes,
               p.name AS part_name, p.article, p.unit,
               sc.cell_code
        FROM movements m
        JOIN parts p ON m.part_id = p.id
        LEFT JOIN storage_cells sc ON sc.id = m.cell_id
        WHERE {where}
        ORDER BY m.created_at DESC
        LIMIT ? OFFSET ?
    ''', params + [per_page, (page - 1) * per_page]).fetchall()

    db.close()

    return render_template('journal.html',
        movements=movements,
        total=total,
        page=page,
        total_pages=max(1, math.ceil(total / per_page)),
        per_page=per_page,
        q=q,
        op_type=op_type,
        date_from=date_from,
        date_to=date_to,
    )


# ── Экспорт в Excel ───────────────────────────────────────────────────────────

def _xl_style(bold=False, bg=None, color='000000', wrap=False, align='left'):
    """Возвращает именованный стиль для ячейки openpyxl."""
    thin = Side(style='thin', color='D0D0D0')
    return {
        'font':      Font(bold=bold, color=color, size=10),
        'fill':      PatternFill('solid', fgColor=bg) if bg else PatternFill(),
        'border':    Border(left=thin, right=thin, top=thin, bottom=thin),
        'alignment': Alignment(wrap_text=wrap, horizontal=align, vertical='center'),
    }


def _apply(cell, style):
    for attr, val in style.items():
        setattr(cell, attr, val)


def _xl_send(wb, filename_prefix):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'{filename_prefix}_{ts}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/journal/export')
@login_required
def journal_export():
    q         = request.args.get('q', '').strip()
    op_type   = request.args.get('op_type', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    db = get_db()
    conditions, params = ['1=1'], []

    if q:
        conditions.append(
            '(p.name LIKE ? OR p.article LIKE ? OR m.operator LIKE ? OR m.engine_number LIKE ?)'
        )
        params += [f'%{q}%'] * 4
    if op_type in ('приход', 'расход'):
        conditions.append('m.operation_type = ?')
        params.append(op_type)
    if date_from:
        conditions.append('DATE(m.created_at) >= ?')
        params.append(date_from)
    if date_to:
        conditions.append('DATE(m.created_at) <= ?')
        params.append(date_to)

    where = ' AND '.join(conditions)
    rows = db.execute(f'''
        SELECT m.created_at, m.operation_type, p.article, p.name, m.quantity, p.unit,
               sc.cell_code, m.engine_number, m.assembly_stage, m.operator, m.notes
        FROM movements m
        JOIN parts p ON m.part_id = p.id
        LEFT JOIN storage_cells sc ON sc.id = m.cell_id
        WHERE {where}
        ORDER BY m.created_at DESC
    ''', params).fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Журнал операций'

    # Заголовки
    headers = [
        '№', 'Дата', 'Время', 'Тип операции', 'Артикул', 'Наименование',
        'Кол-во', 'Ед.изм.', 'Ячейка', 'Двигатель', 'Этап сборки',
        'Исполнитель', 'Примечание',
    ]
    col_w = [5, 12, 8, 13, 18, 32, 8, 7, 11, 16, 22, 22, 28]

    hdr_style = _xl_style(bold=True, bg='1A3A6E', color='FFFFFF', align='center')
    for col, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=col, value=h)
        _apply(cell, hdr_style)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    # Стили строк
    s_in  = _xl_style(bg='EAF7EF')   # приход — светло-зелёный
    s_out = _xl_style(bg='FDEEEE')   # расход — светло-красный
    s_def = _xl_style()

    for idx, row in enumerate(rows, 1):
        dt_str = row[0] or ''
        op     = row[1] or ''
        style  = s_in if op == 'приход' else (s_out if op == 'расход' else s_def)

        values = [
            idx,
            dt_str[:10],
            dt_str[11:16],
            op,
            row[2], row[3],
            row[4], row[5],
            row[6] or '',
            row[7] or '',
            row[8] or '',
            row[9] or '',
            row[10] or '',
        ]
        r = idx + 1
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            _apply(cell, style)
        ws.row_dimensions[r].height = 16

    return _xl_send(wb, 'journal')


@app.route('/export/stock')
@login_required
def stock_export():
    db = get_db()
    rows = db.execute('''
        SELECT p.article, p.name, p.engine_type, p.unit,
               s.quantity, p.red_threshold, p.yellow_threshold,
               p.monthly_plan, sc.cell_code,
               CASE
                 WHEN s.quantity < p.red_threshold   THEN 'Критично'
                 WHEN s.quantity < p.yellow_threshold THEN 'Внимание'
                 ELSE 'Норма'
               END AS status
        FROM parts p
        JOIN stock s  ON s.part_id = p.id
        JOIN storage_cells sc ON sc.id = s.cell_id
        ORDER BY
          CASE WHEN s.quantity < p.red_threshold   THEN 1
               WHEN s.quantity < p.yellow_threshold THEN 2
               ELSE 3 END,
          p.name
    ''').fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Остатки деталей'

    headers = [
        'Артикул', 'Наименование', 'Двигатель', 'Ед.изм.',
        'Остаток', 'Порог красный', 'Порог жёлтый',
        'План/мес', 'Ячейка', 'Статус',
    ]
    col_w = [18, 32, 10, 7, 9, 14, 14, 10, 12, 10]

    hdr_style = _xl_style(bold=True, bg='1A3A6E', color='FFFFFF', align='center')
    for col, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=col, value=h)
        _apply(cell, hdr_style)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    s_red    = _xl_style(bg='FDEEEE')
    s_yellow = _xl_style(bg='FFF9E6')
    s_green  = _xl_style()

    for idx, row in enumerate(rows, 1):
        status = row['status']
        style  = s_red if status == 'Критично' else (s_yellow if status == 'Внимание' else s_green)
        values = [
            row['article'], row['name'], row['engine_type'], row['unit'],
            row['quantity'], row['red_threshold'], row['yellow_threshold'],
            row['monthly_plan'], row['cell_code'], status,
        ]
        r = idx + 1
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            _apply(cell, style)
        ws.row_dimensions[r].height = 16

    return _xl_send(wb, 'stock')


# ── Сборка двигателей ────────────────────────────────────────────────────────

@app.route('/assembly')
@login_required
def assembly():
    db = get_db()
    is_limited = ROLE_LEVEL.get(session.get('role'), 0) < ROLE_LEVEL['master']
    if is_limited:
        base_sql = '''
            SELECT ea.*,
                   COUNT(DISTINCT CASE WHEN ah.is_cancelled=0 THEN ah.id END) AS done_count,
                   (SELECT COUNT(*) FROM assembly_stages
                    WHERE engine_type = ea.engine_type) AS total_stages
            FROM engine_assemblies ea
            LEFT JOIN assembly_history ah ON ah.assembly_id = ea.id
            WHERE ea.assembler = ?
            GROUP BY ea.id ORDER BY ea.started_at DESC'''
        assemblies = db.execute(base_sql, (session['username'],)).fetchall()
    else:
        base_sql = '''
            SELECT ea.*,
                   COUNT(DISTINCT CASE WHEN ah.is_cancelled=0 THEN ah.id END) AS done_count,
                   (SELECT COUNT(*) FROM assembly_stages
                    WHERE engine_type = ea.engine_type) AS total_stages
            FROM engine_assemblies ea
            LEFT JOIN assembly_history ah ON ah.assembly_id = ea.id
            GROUP BY ea.id ORDER BY ea.started_at DESC'''
        assemblies = db.execute(base_sql).fetchall()

    engine_types = db.execute(
        'SELECT DISTINCT engine_type FROM assembly_stages ORDER BY engine_type'
    ).fetchall()
    db.close()
    return render_template('assembly.html',
                           assemblies=assemblies, engine_types=engine_types)


@app.route('/assembly/start', methods=['POST'])
@login_required
def assembly_start():
    engine_number = request.form.get('engine_number', '').strip()
    engine_type   = request.form.get('engine_type', '').strip()
    if not engine_number or not engine_type:
        flash('Укажите номер двигателя и тип.', 'danger')
        return redirect(url_for('assembly'))

    db = get_db()
    if db.execute('SELECT id FROM engine_assemblies WHERE engine_number=?',
                  (engine_number,)).fetchone():
        flash(f'Сборка {engine_number} уже существует.', 'danger')
        db.close()
        return redirect(url_for('assembly'))

    if not db.execute('SELECT COUNT(*) FROM assembly_stages WHERE engine_type=?',
                      (engine_type,)).fetchone()[0]:
        flash(f'Операции для {engine_type} не настроены.', 'danger')
        db.close()
        return redirect(url_for('assembly'))

    cur = db.execute(
        'INSERT INTO engine_assemblies '
        '(engine_number, engine_type, current_stage, status, assembler) '
        'VALUES (?, ?, 1, "В работе", ?)',
        (engine_number, engine_type, session['username'])
    )
    asm_id = cur.lastrowid
    db.commit()
    db.close()
    flash(f'Сборка {engine_number} начата!', 'success')
    return redirect(url_for('assembly_work', asm_id=asm_id))


@app.route('/assembly/<int:asm_id>')
@login_required
def assembly_work(asm_id):
    db = get_db()
    asm = db.execute('SELECT * FROM engine_assemblies WHERE id=?', (asm_id,)).fetchone()
    if not asm:
        flash('Сборка не найдена.', 'danger')
        db.close()
        return redirect(url_for('assembly'))

    if asm['status'] == 'Завершена':
        db.close()
        return redirect(url_for('assembly_finish', asm_id=asm_id))

    all_stages = db.execute(
        'SELECT * FROM assembly_stages WHERE engine_type=? ORDER BY stage_number',
        (asm['engine_type'],)
    ).fetchall()
    total_stages = len(all_stages)

    done_ids = {r['stage_id'] for r in db.execute(
        'SELECT stage_id FROM assembly_history WHERE assembly_id=? AND is_cancelled=0',
        (asm_id,)
    ).fetchall()}

    stages = []
    for s in all_stages:
        if s['id'] in done_ids:
            st = 'done'
        elif s['stage_number'] == asm['current_stage']:
            st = 'current'
        else:
            st = 'pending'
        stages.append({'stage': s, 'status': st})

    current_stage = next((x['stage'] for x in stages if x['status'] == 'current'), None)

    current_parts = []
    if current_stage:
        rows = db.execute('''
            SELECT sp.quantity AS needed,
                   p.id, p.name, p.article, p.unit,
                   p.red_threshold, p.yellow_threshold,
                   s.quantity AS stock_qty, sc.cell_code
            FROM stage_parts sp
            JOIN parts p  ON p.id  = sp.part_id
            JOIN stock s  ON s.part_id = p.id
            JOIN storage_cells sc ON sc.id = s.cell_id
            WHERE sp.stage_id=? ORDER BY p.name
        ''', (current_stage['id'],)).fetchall()

        for r in rows:
            if r['stock_qty'] < r['needed']:
                cs = 'red'
            elif r['stock_qty'] < r['yellow_threshold']:
                cs = 'yellow'
            else:
                cs = 'green'
            current_parts.append({
                'part_id': r['id'], 'name': r['name'],
                'article': r['article'], 'unit': r['unit'],
                'needed': r['needed'], 'stock_qty': r['stock_qty'],
                'cell_code': r['cell_code'], 'card_status': cs,
                'shortage': r['stock_qty'] < r['needed'],
            })

    last_history = db.execute('''
        SELECT ah.id, ah.stage_id, s.stage_name, s.stage_number
        FROM assembly_history ah
        JOIN assembly_stages s ON s.id = ah.stage_id
        WHERE ah.assembly_id=? AND ah.is_cancelled=0
        ORDER BY ah.completed_at DESC LIMIT 1
    ''', (asm_id,)).fetchone()

    done_count   = len(done_ids)
    progress_pct = round(done_count / total_stages * 100) if total_stages else 0
    has_shortage = any(p['shortage'] for p in current_parts)
    db.close()

    return render_template('assembly_work.html',
        asm=asm, stages=stages, current_stage=current_stage,
        current_parts=current_parts, has_shortage=has_shortage,
        total_stages=total_stages, done_count=done_count,
        progress_pct=progress_pct, last_history=last_history)


@app.route('/assembly/<int:asm_id>/complete', methods=['POST'])
@login_required
def assembly_complete(asm_id):
    db = get_db()
    asm = db.execute('SELECT * FROM engine_assemblies WHERE id=?', (asm_id,)).fetchone()
    if not asm or asm['status'] == 'Завершена':
        flash('Сборка не найдена или уже завершена.', 'danger')
        db.close()
        return redirect(url_for('assembly'))

    stage = db.execute(
        'SELECT * FROM assembly_stages WHERE engine_type=? AND stage_number=?',
        (asm['engine_type'], asm['current_stage'])
    ).fetchone()
    if not stage:
        flash('Операция не найдена.', 'danger')
        db.close()
        return redirect(url_for('assembly_work', asm_id=asm_id))

    parts = db.execute('''
        SELECT sp.quantity AS needed, p.id AS part_id, p.name, p.unit,
               s.id AS stock_id, s.quantity AS stock_qty, s.cell_id
        FROM stage_parts sp
        JOIN parts p ON p.id = sp.part_id
        JOIN stock s ON s.part_id = p.id
        WHERE sp.stage_id=?
    ''', (stage['id'],)).fetchall()

    shortages = [p for p in parts if p['stock_qty'] < p['needed']]
    if shortages:
        names = ', '.join(
            f'{p["name"]} (нужно {p["needed"]}, есть {p["stock_qty"]})' for p in shortages
        )
        flash(f'Недостаточно деталей: {names}', 'danger')
        db.close()
        return redirect(url_for('assembly_work', asm_id=asm_id))

    for p in parts:
        db.execute(
            'UPDATE stock SET quantity=quantity-?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
            (p['needed'], p['stock_id'])
        )
        db.execute(
            'INSERT INTO movements (part_id, cell_id, operation_type, quantity, '
            'engine_number, assembly_stage, operator) VALUES (?,?,?,?,?,?,?)',
            (p['part_id'], p['cell_id'], 'расход', p['needed'],
             asm['engine_number'], stage['stage_name'], session['full_name'])
        )

    db.execute(
        'INSERT INTO assembly_history (assembly_id, stage_id, assembler) VALUES (?,?,?)',
        (asm_id, stage['id'], session['username'])
    )

    total_stages = db.execute(
        'SELECT COUNT(*) FROM assembly_stages WHERE engine_type=?', (asm['engine_type'],)
    ).fetchone()[0]
    next_stage = asm['current_stage'] + 1

    if next_stage > total_stages:
        db.execute(
            'UPDATE engine_assemblies SET current_stage=?, status="Завершена", '
            'completed_at=CURRENT_TIMESTAMP WHERE id=?',
            (next_stage, asm_id)
        )
        db.commit()
        db.close()
        flash(f'Сборка {asm["engine_number"]} завершена!', 'success')
        return redirect(url_for('assembly_finish', asm_id=asm_id))
    else:
        db.execute('UPDATE engine_assemblies SET current_stage=? WHERE id=?',
                   (next_stage, asm_id))
        db.commit()
        db.close()
        flash(f'Операция «{stage["stage_name"]}» выполнена. Переход к {next_stage}-й.', 'success')
        return redirect(url_for('assembly_work', asm_id=asm_id))


@app.route('/assembly/<int:asm_id>/cancel', methods=['POST'])
@login_required
def assembly_cancel(asm_id):
    db = get_db()
    asm = db.execute('SELECT * FROM engine_assemblies WHERE id=?', (asm_id,)).fetchone()
    if not asm or asm['status'] == 'Завершена':
        flash('Нельзя отменить завершённую сборку.', 'danger')
        db.close()
        return redirect(url_for('assembly'))

    last = db.execute('''
        SELECT ah.id, ah.stage_id, s.stage_name, s.stage_number
        FROM assembly_history ah
        JOIN assembly_stages s ON s.id = ah.stage_id
        WHERE ah.assembly_id=? AND ah.is_cancelled=0
        ORDER BY ah.completed_at DESC LIMIT 1
    ''', (asm_id,)).fetchone()

    if not last:
        flash('Нет выполненных операций для отмены.', 'warning')
        db.close()
        return redirect(url_for('assembly_work', asm_id=asm_id))

    parts = db.execute('''
        SELECT sp.quantity AS needed, p.id AS part_id, s.id AS stock_id, s.cell_id
        FROM stage_parts sp
        JOIN parts p ON p.id = sp.part_id
        JOIN stock s ON s.part_id = p.id
        WHERE sp.stage_id=?
    ''', (last['stage_id'],)).fetchall()

    for p in parts:
        db.execute(
            'UPDATE stock SET quantity=quantity+?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
            (p['needed'], p['stock_id'])
        )
        db.execute(
            'INSERT INTO movements (part_id, cell_id, operation_type, quantity, '
            'engine_number, assembly_stage, operator, notes) VALUES (?,?,?,?,?,?,?,?)',
            (p['part_id'], p['cell_id'], 'приход', p['needed'],
             asm['engine_number'], last['stage_name'], session['full_name'],
             f'Отмена операции «{last["stage_name"]}»')
        )

    db.execute('UPDATE assembly_history SET is_cancelled=1 WHERE id=?', (last['id'],))
    db.execute('UPDATE engine_assemblies SET current_stage=?, status="В работе" WHERE id=?',
               (last['stage_number'], asm_id))
    db.commit()
    db.close()
    flash(f'Операция «{last["stage_name"]}» отменена. Детали возвращены.', 'warning')
    return redirect(url_for('assembly_work', asm_id=asm_id))


@app.route('/assembly/<int:asm_id>/finish')
@login_required
def assembly_finish(asm_id):
    db = get_db()
    asm = db.execute('SELECT * FROM engine_assemblies WHERE id=?', (asm_id,)).fetchone()
    if not asm:
        flash('Сборка не найдена.', 'danger')
        db.close()
        return redirect(url_for('assembly'))

    all_stages = db.execute(
        'SELECT * FROM assembly_stages WHERE engine_type=? ORDER BY stage_number',
        (asm['engine_type'],)
    ).fetchall()
    total_stages = len(all_stages)

    hist_rows = db.execute('''
        SELECT ah.stage_id, ah.completed_at, ah.assembler
        FROM assembly_history ah
        WHERE ah.assembly_id=? AND ah.is_cancelled=0
        ORDER BY ah.completed_at
    ''', (asm_id,)).fetchall()
    hist_by_sid = {r['stage_id']: r for r in hist_rows}
    done_ids = set(hist_by_sid.keys())

    stages = []
    for s in all_stages:
        if s['id'] in done_ids:
            st = 'done'
        elif s['stage_number'] == asm['current_stage']:
            st = 'current'
        else:
            st = 'pending'
        stages.append({'stage': s, 'status': st, 'history': hist_by_sid.get(s['id'])})

    done_count   = len(done_ids)
    progress_pct = round(done_count / total_stages * 100) if total_stages else 0

    db.close()
    return render_template('assembly_finish.html',
        asm=asm, stages=stages, total_stages=total_stages,
        done_count=done_count, progress_pct=progress_pct)


# ── Настройки (только admin) ──────────────────────────────────────────────────

def _admin_required():
    return ROLE_LEVEL.get(session.get('role'), 0) >= ROLE_LEVEL['admin']


@app.route('/settings')
@login_required
def settings():
    if not _admin_required():
        flash('Доступ только для администратора.', 'danger')
        return redirect(url_for('dashboard'))
    db  = get_db()
    tab = request.args.get('tab', 'users')

    users      = db.execute("SELECT id, username, full_name, role FROM users ORDER BY role, username").fetchall()
    all_parts  = db.execute(
        "SELECT id, article, name, red_threshold, yellow_threshold FROM parts ORDER BY id"
    ).fetchall()
    stages     = db.execute(
        "SELECT * FROM assembly_stages ORDER BY engine_type, stage_number"
    ).fetchall()
    eng_types  = [r['engine_type'] for r in
                  db.execute("SELECT DISTINCT engine_type FROM assembly_stages ORDER BY engine_type").fetchall()]
    db.close()
    return render_template('settings.html',
        tab=tab, users=users, all_parts=all_parts,
        stages=stages, eng_types=eng_types,
        ROLE_NAMES=ROLE_NAMES)


@app.route('/settings/users/add', methods=['POST'])
@login_required
def settings_users_add():
    if not _admin_required():
        return redirect(url_for('dashboard'))
    from werkzeug.security import generate_password_hash
    username  = request.form['username'].strip()
    password  = request.form['password']
    full_name = request.form['full_name'].strip()
    role      = request.form['role']
    if not username or not password:
        flash('Логин и пароль обязательны.', 'danger')
        return redirect(url_for('settings', tab='users'))
    db = get_db()
    try:
        db.execute(
            "INSERT INTO users (username, password_hash, full_name, role) VALUES (?,?,?,?)",
            (username, generate_password_hash(password), full_name, role)
        )
        db.commit()
        flash(f'Пользователь «{username}» создан.', 'success')
    except Exception:
        flash('Логин уже занят.', 'danger')
    db.close()
    return redirect(url_for('settings', tab='users'))


@app.route('/settings/users/<int:uid>/edit', methods=['POST'])
@login_required
def settings_users_edit(uid):
    if not _admin_required():
        return redirect(url_for('dashboard'))
    from werkzeug.security import generate_password_hash
    full_name = request.form['full_name'].strip()
    role      = request.form['role']
    new_pass  = request.form.get('new_password', '').strip()
    db = get_db()
    if new_pass:
        db.execute(
            "UPDATE users SET full_name=?, role=?, password_hash=? WHERE id=?",
            (full_name, role, generate_password_hash(new_pass), uid)
        )
    else:
        db.execute("UPDATE users SET full_name=?, role=? WHERE id=?", (full_name, role, uid))
    db.commit()
    db.close()
    flash('Пользователь обновлён.', 'success')
    return redirect(url_for('settings', tab='users'))


@app.route('/settings/users/<int:uid>/delete', methods=['POST'])
@login_required
def settings_users_delete(uid):
    if not _admin_required():
        return redirect(url_for('dashboard'))
    if uid == session['user_id']:
        flash('Нельзя удалить самого себя.', 'danger')
        return redirect(url_for('settings', tab='users'))
    db = get_db()
    db.execute("DELETE FROM users WHERE id=?", (uid,))
    db.commit()
    db.close()
    flash('Пользователь удалён.', 'success')
    return redirect(url_for('settings', tab='users'))


@app.route('/settings/thresholds', methods=['POST'])
@login_required
def settings_thresholds():
    if not _admin_required():
        return redirect(url_for('dashboard'))
    db = get_db()
    part_ids = request.form.getlist('part_id')
    updated = 0
    for pid in part_ids:
        red = request.form.get(f'red_{pid}', type=int)
        yel = request.form.get(f'yel_{pid}', type=int)
        if red is not None and yel is not None and yel >= red >= 0:
            db.execute(
                "UPDATE parts SET red_threshold=?, yellow_threshold=? WHERE id=?",
                (red, yel, int(pid))
            )
            updated += 1
    db.commit()
    db.close()
    flash(f'Обновлено порогов: {updated}.', 'success')
    return redirect(url_for('settings', tab='thresholds'))


@app.route('/settings/stages/add', methods=['POST'])
@login_required
def settings_stages_add():
    if not _admin_required():
        return redirect(url_for('dashboard'))
    engine_type  = request.form['engine_type'].strip()
    stage_number = request.form.get('stage_number', type=int)
    stage_name   = request.form['stage_name'].strip()
    description  = request.form.get('description', '').strip()
    if not engine_type or not stage_name or not stage_number:
        flash('Заполните все обязательные поля.', 'danger')
        return redirect(url_for('settings', tab='stages'))
    db = get_db()
    db.execute(
        "INSERT INTO assembly_stages (engine_type, stage_number, stage_name, description) VALUES (?,?,?,?)",
        (engine_type, stage_number, stage_name, description or None)
    )
    db.commit()
    db.close()
    flash(f'Операция «{stage_name}» добавлена.', 'success')
    return redirect(url_for('settings', tab='stages'))


@app.route('/settings/stages/<int:sid>/delete', methods=['POST'])
@login_required
def settings_stages_delete(sid):
    if not _admin_required():
        return redirect(url_for('dashboard'))
    db = get_db()
    used = db.execute(
        "SELECT COUNT(*) FROM assembly_history ah WHERE ah.stage_id=? AND ah.is_cancelled=0",
        (sid,)
    ).fetchone()[0]
    if used:
        flash('Нельзя удалить операцию — есть история выполнения.', 'danger')
    else:
        db.execute("DELETE FROM stage_parts WHERE stage_id=?", (sid,))
        db.execute("DELETE FROM assembly_stages WHERE id=?", (sid,))
        db.commit()
        flash('Операция удалена.', 'success')
    db.close()
    return redirect(url_for('settings', tab='stages'))


# ── Отчёты ────────────────────────────────────────────────────────────────────

@app.route('/reports')
@login_required
def reports():
    if ROLE_LEVEL.get(session.get('role'), 0) < ROLE_LEVEL['master']:
        flash('Недостаточно прав для просмотра отчётов.', 'danger')
        return redirect(url_for('dashboard'))

    db  = get_db()
    tab = request.args.get('tab', 'spending')

    # ── Вкладка 1: Расход за период ──────────────────────────────────────────
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to',   '')
    part_id   = request.args.get('part_id',   '', type=str)
    engine_q  = request.args.get('engine_q',  '').strip()

    if not date_from:
        date_from = db.execute("SELECT date('now', '-29 days')").fetchone()[0]
    if not date_to:
        date_to = db.execute("SELECT date('now')").fetchone()[0]

    spend_sql = """
        SELECT date(m.created_at) AS day,
               SUM(m.quantity)    AS total_qty,
               COUNT(*)           AS ops
        FROM movements m
        WHERE m.operation_type = 'расход'
          AND date(m.created_at) BETWEEN ? AND ?
    """
    spend_params = [date_from, date_to]
    if part_id:
        spend_sql   += " AND m.part_id = ?"
        spend_params.append(int(part_id))
    if engine_q:
        spend_sql   += " AND m.engine_number LIKE ?"
        spend_params.append(f'%{engine_q}%')
    spend_sql += " GROUP BY day ORDER BY day"

    spend_rows = db.execute(spend_sql, spend_params).fetchall()
    spend_days = [r['day']       for r in spend_rows]
    spend_vals = [r['total_qty'] for r in spend_rows]

    # Детальная таблица расхода
    detail_sql = """
        SELECT m.created_at, p.name AS part_name, p.article,
               m.quantity, m.engine_number, m.assembly_stage, m.operator
        FROM movements m
        JOIN parts p ON p.id = m.part_id
        WHERE m.operation_type = 'расход'
          AND date(m.created_at) BETWEEN ? AND ?
    """
    detail_params = [date_from, date_to]
    if part_id:
        detail_sql   += " AND m.part_id = ?"
        detail_params.append(int(part_id))
    if engine_q:
        detail_sql   += " AND m.engine_number LIKE ?"
        detail_params.append(f'%{engine_q}%')
    detail_sql += " ORDER BY m.created_at DESC LIMIT 200"
    spend_detail = db.execute(detail_sql, detail_params).fetchall()

    all_parts = db.execute(
        "SELECT id, name, article FROM parts ORDER BY name"
    ).fetchall()

    # ── Вкладка 2: Сборки ────────────────────────────────────────────────────
    assemblies = db.execute("""
        SELECT ea.id, ea.engine_number, ea.engine_type,
               ea.current_stage, ea.status, ea.assembler,
               ea.started_at, ea.completed_at,
               (SELECT COUNT(*) FROM assembly_stages WHERE engine_type=ea.engine_type) AS total_stages,
               (SELECT COUNT(*) FROM assembly_history ah
                WHERE ah.assembly_id=ea.id AND ah.is_cancelled=0) AS done_count
        FROM engine_assemblies ea
        ORDER BY ea.started_at DESC
    """).fetchall()

    asm_list = []
    for a in assemblies:
        pct = round(a['done_count'] / a['total_stages'] * 100) if a['total_stages'] else 0
        asm_list.append({**dict(a), 'pct': pct})

    # ── Вкладка 3: Дефицит ───────────────────────────────────────────────────
    deficit = db.execute("""
        SELECT p.id, p.article, p.name, p.engine_type, p.unit,
               p.red_threshold, p.yellow_threshold, p.monthly_plan,
               COALESCE(s.quantity, 0) AS stock_qty,
               sc.cell_code
        FROM parts p
        LEFT JOIN stock s  ON s.part_id = p.id
        LEFT JOIN storage_cells sc ON sc.id = s.cell_id
        WHERE COALESCE(s.quantity, 0) < COALESCE(p.red_threshold, 0)
        ORDER BY (p.red_threshold - COALESCE(s.quantity, 0)) DESC
    """).fetchall()

    # Кол-во позиций по зонам
    zone_counts = db.execute("""
        SELECT
            SUM(CASE WHEN COALESCE(s.quantity,0) < COALESCE(p.red_threshold,0)    THEN 1 ELSE 0 END) AS red_cnt,
            SUM(CASE WHEN COALESCE(s.quantity,0) >= COALESCE(p.red_threshold,0)
                      AND COALESCE(s.quantity,0) < COALESCE(p.yellow_threshold,0) THEN 1 ELSE 0 END) AS yellow_cnt,
            SUM(CASE WHEN COALESCE(s.quantity,0) >= COALESCE(p.yellow_threshold,0) THEN 1 ELSE 0 END) AS green_cnt
        FROM parts p LEFT JOIN stock s ON s.part_id = p.id
    """).fetchone()

    db.close()

    import json
    return render_template('reports.html',
        tab=tab,
        date_from=date_from, date_to=date_to,
        part_id=part_id, engine_q=engine_q,
        spend_days_json=json.dumps(spend_days, ensure_ascii=False),
        spend_vals_json=json.dumps(spend_vals),
        spend_detail=spend_detail,
        spend_total=sum(spend_vals),
        all_parts=all_parts,
        asm_list=asm_list,
        deficit=deficit,
        zone_counts=zone_counts,
    )


# ── Ячейки хранения ───────────────────────────────────────────────────────────

@app.route('/cells')
@login_required
def cells():
    db  = get_db()
    q   = request.args.get('q', '').strip()

    sql = '''
        SELECT sc.id, sc.cell_code, sc.shelf_number, sc.level_number, sc.cell_number,
               p.id AS part_id, p.name AS part_name, p.article,
               p.red_threshold, p.yellow_threshold, p.unit,
               COALESCE(s.quantity, 0) AS quantity
        FROM storage_cells sc
        LEFT JOIN stock s  ON s.cell_id  = sc.id
        LEFT JOIN parts p  ON p.id = s.part_id
        WHERE 1=1
    '''
    params = []
    if q:
        sql += " AND (sc.cell_code LIKE ? OR p.name LIKE ? OR p.article LIKE ?)"
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    sql += " ORDER BY sc.shelf_number, sc.level_number, sc.cell_number"

    rows = db.execute(sql, params).fetchall()
    db.close()

    def cell_status(r):
        if not r['part_id']:
            return 'empty'
        qty, red, yel = r['quantity'], r['red_threshold'] or 0, r['yellow_threshold'] or 0
        if qty < red:
            return 'red'
        if qty < yel:
            return 'yellow'
        return 'green'

    cells_list = []
    for r in rows:
        cells_list.append({**dict(r), 'status': cell_status(r)})

    # Группировка для схемы: {shelf: {level: [cell, ...]}}
    grid = {}
    for c in cells_list:
        sh = c['shelf_number']
        lv = c['level_number']
        grid.setdefault(sh, {}).setdefault(lv, []).append(c)

    shelves     = sorted(grid.keys())
    all_levels  = sorted({lv for sh in grid.values() for lv in sh})

    stats = {
        'total': len(cells_list),
        'empty':  sum(1 for c in cells_list if c['status'] == 'empty'),
        'red':    sum(1 for c in cells_list if c['status'] == 'red'),
        'yellow': sum(1 for c in cells_list if c['status'] == 'yellow'),
        'green':  sum(1 for c in cells_list if c['status'] == 'green'),
    }

    return render_template('cells.html',
                           cells=cells_list, q=q,
                           grid=grid, shelves=shelves, all_levels=all_levels,
                           stats=stats)


# ── Справочник деталей ────────────────────────────────────────────────────────

@app.route('/details')
@login_required
def details():
    db  = get_db()
    q   = request.args.get('q', '').strip()
    eng = request.args.get('engine', '')

    sql = '''
        SELECT p.id, p.article, p.name, p.engine_type, p.unit,
               p.monthly_plan, p.red_threshold, p.yellow_threshold,
               s.quantity AS stock_qty,
               sc.cell_code
        FROM parts p
        LEFT JOIN stock s  ON s.part_id = p.id
        LEFT JOIN storage_cells sc ON sc.id = s.cell_id
        WHERE 1=1
    '''
    params = []
    if q:
        sql += " AND (p.name LIKE ? OR p.article LIKE ?)"
        params += [f'%{q}%', f'%{q}%']
    if eng:
        sql += " AND p.engine_type = ?"
        params.append(eng)
    sql += " ORDER BY p.id"

    rows = db.execute(sql, params).fetchall()

    engine_types = [r['engine_type'] for r in
                    db.execute("SELECT DISTINCT engine_type FROM parts ORDER BY engine_type").fetchall()]

    parts = []
    for r in rows:
        qty   = r['stock_qty'] or 0
        red   = r['red_threshold'] or 0
        yel   = r['yellow_threshold'] or 0
        if qty < red:
            status = 'red'
        elif qty < yel:
            status = 'yellow'
        else:
            status = 'green'
        parts.append({**dict(r), 'stock_qty': qty, 'status': status})

    db.close()
    return render_template('details.html',
                           parts=parts, q=q, engine=eng,
                           engine_types=engine_types)


@app.route('/details/<int:part_id>/thresholds', methods=['POST'])
@login_required
def details_thresholds(part_id):
    if ROLE_LEVEL.get(session.get('role'), 0) < ROLE_LEVEL['master']:
        flash('Недостаточно прав.', 'danger')
        return redirect(url_for('details'))
    red = request.form.get('red_threshold', type=int)
    yel = request.form.get('yellow_threshold', type=int)
    if red is None or yel is None or red < 0 or yel < red:
        flash('Некорректные значения порогов.', 'danger')
        return redirect(url_for('details'))
    db = get_db()
    db.execute(
        'UPDATE parts SET red_threshold=?, yellow_threshold=? WHERE id=?',
        (red, yel, part_id)
    )
    db.commit()
    db.close()
    flash('Пороги обновлены.', 'success')
    return redirect(url_for('details') + (f'?q={request.form.get("back_q","")}'
                                          if request.form.get('back_q') else ''))


# ── Запуск ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'csupd.db')
    if not os.path.exists(db_path):
        from init_db import main as init_db
        init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
